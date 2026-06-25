from dataclasses import dataclass
from numbers import Integral, Real
from pathlib import Path
from typing import Iterator

from django.contrib.auth import get_user_model
from django.core.exceptions import ValidationError
from django.core.validators import validate_email
from django.db import IntegrityError, transaction
from openpyxl import load_workbook

from marketing.models import MailingMessage
from marketing.tasks import send_mailing_message

REQUIRED_HEADERS = ("external_id", "user_id", "email", "subject", "message")


class MailingImportError(Exception):
    pass


class RowValidationError(MailingImportError):
    pass


@dataclass(frozen=True)
class MailingRow:
    external_id: str
    user_id: int
    email: str
    subject: str
    message: str


@dataclass
class ImportStats:
    processed_rows: int = 0
    created_records: int = 0
    skipped_records: int = 0
    error_rows: int = 0


def import_mailing_messages(
    file_path: str | Path,
    *,
    sheet_name: str | None = None,
) -> ImportStats:
    stats = ImportStats()

    for _row_number, row_data in iter_mailing_rows(file_path, sheet_name=sheet_name):
        stats.processed_rows += 1
        try:
            mailing_row = validate_mailing_row(row_data)
            created = create_mailing_message(mailing_row)
        except RowValidationError:
            stats.error_rows += 1
            continue

        if created:
            stats.created_records += 1
        else:
            stats.skipped_records += 1

    return stats


def iter_mailing_rows(
    file_path: str | Path,
    *,
    sheet_name: str | None = None,
) -> Iterator[tuple[int, dict[str, object]]]:
    workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
    try:
        try:
            worksheet = workbook[sheet_name] if sheet_name else workbook.active
        except KeyError as exc:
            raise MailingImportError(f"Worksheet '{sheet_name}' does not exist.") from exc
        rows = worksheet.iter_rows(values_only=True)

        try:
            header_row = next(rows)
        except StopIteration as exc:
            raise MailingImportError("XLSX file is empty.") from exc

        headers = [str(value).strip() if value is not None else "" for value in header_row]
        missing_headers = [header for header in REQUIRED_HEADERS if header not in headers]
        if missing_headers:
            raise MailingImportError(f"Missing required columns: {', '.join(missing_headers)}.")

        header_indexes = {header: headers.index(header) for header in REQUIRED_HEADERS}

        for row_number, row in enumerate(rows, start=2):
            if all(value is None or str(value).strip() == "" for value in row):
                continue
            yield row_number, {
                header: row[index] if index < len(row) else None
                for header, index in header_indexes.items()
            }
    finally:
        workbook.close()


def validate_mailing_row(row_data: dict[str, object]) -> MailingRow:
    external_id = _to_required_string(row_data.get("external_id"), "external_id")
    user_id = _to_user_id(row_data.get("user_id"))
    email = _to_required_string(row_data.get("email"), "email")
    subject = _to_required_string(row_data.get("subject"), "subject")
    message = _to_required_string(row_data.get("message"), "message")

    if len(external_id) > 255:
        raise RowValidationError("external_id is too long.")
    if len(subject) > 255:
        raise RowValidationError("subject is too long.")

    try:
        validate_email(email)
    except ValidationError as exc:
        raise RowValidationError("email is invalid.") from exc

    return MailingRow(
        external_id=external_id,
        user_id=user_id,
        email=email,
        subject=subject,
        message=message,
    )


def create_mailing_message(mailing_row: MailingRow) -> bool:
    user_model = get_user_model()
    try:
        user = user_model.objects.get(pk=mailing_row.user_id)
    except user_model.DoesNotExist as exc:
        raise RowValidationError("user_id does not exist.") from exc

    try:
        with transaction.atomic():
            mailing_message = MailingMessage.objects.create(
                external_id=mailing_row.external_id,
                user=user,
                email=mailing_row.email,
                subject=mailing_row.subject,
                message=mailing_row.message,
            )
            transaction.on_commit(
                lambda message_id=mailing_message.pk: send_mailing_message.apply_async(
                    args=[message_id],
                )
            )
    except IntegrityError:
        return False

    return True


def _to_required_string(value: object, field_name: str) -> str:
    if value is None:
        raise RowValidationError(f"{field_name} is required.")

    normalized = str(value).strip()
    if not normalized:
        raise RowValidationError(f"{field_name} is required.")

    return normalized


def _to_user_id(value: object) -> int:
    if value is None:
        raise RowValidationError("user_id is required.")

    if isinstance(value, Integral):
        return int(value)

    if isinstance(value, Real):
        if value.is_integer():
            return int(value)
        raise RowValidationError("user_id must be an integer.")

    try:
        return int(value)
    except (TypeError, ValueError) as exc:
        raise RowValidationError("user_id must be an integer.") from exc
